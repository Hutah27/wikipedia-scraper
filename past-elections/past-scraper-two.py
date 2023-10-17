import os
import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
import re

# Function to process a Wikipedia page
def process_wikipedia_page(page_name):
    # Display a message that the script is processing the page
    print(f"\x1b[33mProcessing {page_name}...\x1b[0m")

    # Construct the URL for the Wikipedia page
    page_url = f"https://en.wikipedia.org/wiki/{page_name.replace(' ', '_')}"

    # Send an HTTP GET request to the URL
    response = requests.get(page_url)

    # Check if the request was successful
    if response.status_code != 200:
        print(f"\x1b[31mFailed to retrieve data for page: {page_name}\x1b[0m") 
        return

    # Create a BeautifulSoup object to parse the HTML content
    soup = BeautifulSoup(response.text, "html.parser")

    # Check if the page contains a table with the specified class
    table = soup.find("table", {"class": "wikitable plainrowheaders"})
    if table:
        print(f"\x1b[31mPlease use past-scraper-one.py to extract data from {page_name}.\x1b[0m") 
        return

    # Get the title of the page and use it as the directory name
    page_title = soup.find("title").get_text()
    page_title = page_title.split(" - Wikipedia")[0]
    output_directory = page_title.replace(" ", "_")

    # Create a new directory for the page if it doesn't exist
    if not os.path.exists(output_directory):
        os.mkdir(output_directory)

    # Check if the page has a section with the specified name
    section_name = "Election results"  # Change this to the section name you want to check
    section = soup.find("span", {"id": section_name})
    if section:
        # Check if the section has already been processed
        if section_name in processed_sections:
            return

        # Add the processed section and directory to the dictionary
        processed_sections[section_name] = output_directory

        # Display the message
        print(f"\x1b[32mData saved for '{section_name}'\x1b[0m") 

    # Find all section headings (e.g., <span class="mw-headline">Section Name</span>)
    section_headings = soup.find_all("span", class_="mw-headline")

    # Create an empty list to store data
    all_data = []

    # Iterate through the section headings
    for index, section_heading in enumerate(section_headings, 1):
        section_name = section_heading.get_text()

        # Check if the section should be skipped
        if section_name in ["See also", "Notes", "References", "Bibliography", "External links"]:
            continue

        # Find the table associated with the section
        table = section_heading.find_next("table", class_="wikitable")

        # Check if the table has the class "plainrowheaders"
        if table and "plainrowheaders" in table.get("class", []):
            continue

        data = {
            "District": [],
            "Representative": [],
            "Party": [],
            "First Elected": [],
            "Results": [],
            "Candidates": [],
        }

        if table:
            rows = table.find_all("tr")
            for row in rows[1:]:  # Skip the header row
                columns = row.find_all(["th", "td"])
                if len(columns) == 6:
                    data["District"].append(columns[0].get_text(strip=True))
                    data["Representative"].append(columns[1].get_text(strip=True))
                    data["Party"].append(columns[2].get_text(strip=True))
                    data["First Elected"].append(columns[3].get_text(strip=True))

                    # Replace line breaks with spaces in Excel cells
                    results_text = columns[4].get_text().replace("\n", " ")
                    candidates_text = columns[5].get_text().replace("\n", " ")

                    data["Results"].append(results_text)
                    data["Candidates"].append(candidates_text)

        all_data.append((section_name, data))

    # Create an Excel file for each section and save in the directory
    for section_name, data in all_data:
        df = pd.DataFrame(data)

        # Create a new Excel workbook
        wb = Workbook()
        ws = wb.active

        # Add the data to the Excel sheet
        for row in dataframe_to_rows(df, index=False, header=True):
            ws.append(row)

        # Adjust column widths and row heights
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[get_column_letter(column[0].column)].width = adjusted_width

        for row in ws.iter_rows():
            max_height = 0
            for cell in row:
                if cell.value:
                    lines = cell.value.count(" ") + 1
                    adjusted_height = 12 * lines
                    if adjusted_height > max_height:
                        max_height = adjusted_height
            ws.row_dimensions[row[0].row].height = max_height

        # Save the Excel workbook in the directory with the section name
        section_name = section_name.replace(" ", "_").replace("/", "_").replace(":", "_").replace("(", "_").replace(")", "_")
        output_file = os.path.join(output_directory, f"{section_name}.xlsx")
        wb.save(output_file)

        # Display a message indicating that data and files are saved
        print(f"\x1b[32mData saved for {section_name}\x1b[0m")

if __name__ == "__main__":
    # Input one or more Wikipedia page names
    print("Enter Wikipedia page names one per line. Type 'done' on a new line when finished")
    page_names = []

    while True:
        page_name = input()
        if page_name.lower() == "done":
            break
        page_names.append(page_name)

    # Process each page name
    for page_name in page_names:
        process_wikipedia_page(page_name.strip())

    print("\nThank you for using the Wikipedia scraper for past elections.")
    print("Author: Gurwinder Singh")
    print("Version: 1.0")
    print("For any issues, please contact on Upwork: freelancers/~0162de9053b9e180f4")
